# -*- coding: utf-8 -*-
"""
Genera BANQUEO_GUYTON.xlsx a partir del Excel de simulacros UNA Puno, tomando
SOLO los cursos "de letras" (conceptuales, sin cálculo numérico) que el usuario
pidió: excluye Aritmética, Álgebra, Geometría, Trigonometría, Física, Química y
Razonamiento Matemático. Incluye Biología y Anatomía (ciencia pero teórica).

Salida (una sola hoja para subir como pestaña a BD Guyton):
  - banqueo_preguntas: todas las preguntas normalizadas, con columna `curso`.
  - banqueo_progreso : hoja vacía (solo headers) para la gamificación simple.

IMPORTANTE: NO copia ninguna hoja con datos personales (usuarios, historial,
confirmado, registros). Solo preguntas. El .xlsx resultante está en .gitignore
(material de la academia); se sube manualmente al Sheet privado BD Guyton.

Uso:  py -3.12 docs/generar_banqueo.py
"""

import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

SRC = r"C:\PROGRAMACION\ACADEMIA GUYTON\UNA PUNO SIMULACROS (1).xlsx"
OUT = r"C:\PROGRAMACION\ACADEMIA GUYTON\BANQUEO_GUYTON.xlsx"

# Hojas de letras a incluir (nombre de hoja -> nombre de curso visible)
CURSOS = {
    "Banco_Comunicación": "Comunicación",
    "Banco_Literatura": "Literatura",
    "Banco_Razonamiento Verbal": "Razonamiento Verbal",
    "Banco_Psicología y Filosofía": "Psicología y Filosofía",
    "Banco_Historia": "Historia",
    "Banco_Geografía": "Geografía",
    "Banco_Economía": "Economía",
    "Banco_Educación Cívica": "Educación Cívica",
    "Banco_Biología y Anatomía": "Biología y Anatomía",
    "Banco_Inglés": "Inglés",
    "Banco_Quechua y aimara": "Quechua y Aimara",
}

HEADERS = [
    "id_pregunta", "curso", "tema", "subtema", "enunciado",
    "opcion_1", "opcion_2", "opcion_3", "opcion_4", "opcion_5",
    "correcta", "justificacion", "tiempo_seg", "imagen_url", "fuente", "estado",
]

# Palabras que delatan que la pregunta necesita una figura que NO tenemos.
PISTAS_IMAGEN = ["gráfic", "grafic", "figura", "en la imagen", "adjunt",
                 "mostrado", "se muestra", "siguiente esquema"]

HEADER_FILL = PatternFill(start_color="050B2B", end_color="050B2B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def limpio(v):
    return "" if v is None else str(v).strip()


def main():
    if not os.path.exists(SRC):
        raise SystemExit("No se encontró el Excel de origen:\n  " + SRC)
    wb = load_workbook(SRC, read_only=True, data_only=True)

    filas = []
    n = 0
    stats = {}
    borrador_por_imagen = 0
    descartadas = 0

    for hoja, curso in CURSOS.items():
        if hoja not in wb.sheetnames:
            print("AVISO: falta la hoja", hoja); continue
        ws = wb[hoja]
        ok = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # headers
            row = list(row) + [""] * (16 - len(row))
            enunciado = limpio(row[0])
            if not enunciado:
                continue
            opciones = [limpio(row[2]), limpio(row[3]), limpio(row[4]), limpio(row[5]), limpio(row[6])]
            n_op = sum(1 for o in opciones if o)
            # correcta: viene como número 1..5 (a veces float "5.0")
            try:
                correcta = int(float(limpio(row[7])))
            except (ValueError, TypeError):
                correcta = 0
            # validaciones mínimas: >=2 opciones y una correcta existente y no vacía
            if n_op < 2 or correcta < 1 or correcta > 5 or not opciones[correcta - 1]:
                descartadas += 1
                continue
            try:
                tiempo = int(float(limpio(row[8]))) or 90
            except (ValueError, TypeError):
                tiempo = 90
            imagen = limpio(row[9])
            # Si parece necesitar figura y no hay imagen -> se deja en borrador
            estado = "publicado"
            low = enunciado.lower()
            if not imagen and any(p in low for p in PISTAS_IMAGEN):
                estado = "borrador"
                borrador_por_imagen += 1

            n += 1
            filas.append([
                "bpg-%05d" % n, curso, limpio(row[13]), limpio(row[14]), enunciado,
                opciones[0], opciones[1], opciones[2], opciones[3], opciones[4],
                correcta, limpio(row[10]), tiempo, imagen, limpio(row[15]), estado,
            ])
            ok += 1
        stats[curso] = ok

    # --- Escribir salida ---
    out = Workbook()
    ws = out.active
    ws.title = "banqueo_preguntas"
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=c).fill = HEADER_FILL
        ws.cell(row=1, column=c).font = HEADER_FONT
    for f in filas:
        ws.append(f)
    ws.freeze_panes = "A2"

    prog = out.create_sheet("banqueo_progreso")
    PROG_HEADERS = ["id_progreso", "id_usuario", "curso", "respondidas", "correctas", "ultima_practica"]
    prog.append(PROG_HEADERS)
    for c in range(1, len(PROG_HEADERS) + 1):
        prog.cell(row=1, column=c).fill = HEADER_FILL
        prog.cell(row=1, column=c).font = HEADER_FONT
    prog.freeze_panes = "A2"

    out.save(OUT)

    print("Generado:", OUT)
    print("Total preguntas válidas:", n)
    print("  Descartadas (sin opciones / correcta inválida):", descartadas)
    print("  En 'borrador' por requerir figura sin imagen:", borrador_por_imagen)
    print("Por curso:")
    for curso, c in sorted(stats.items(), key=lambda kv: -kv[1]):
        print("  %-24s %d" % (curso, c))


if __name__ == "__main__":
    main()
