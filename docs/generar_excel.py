# -*- coding: utf-8 -*-
"""
Genera BASE_DATOS_GUYTON.xlsx en la raíz del proyecto ACADEMIA GUYTON.

Crea la hoja LEEME + las 12 hojas del modelo de datos descrito en
docs/MODELO_DATOS.md, con headers formateados, freeze de fila 1, anchos de
columna razonables, y 2-4 filas de ejemplo por hoja marcadas como DEMO
(ids con el patrón "-demo-N", y la palabra DEMO en observaciones/descripcion
donde aplica) y coherentes entre sí (mismos FKs cruzados).

Uso:
    py -3.12 docs/generar_excel.py

(En este equipo el "python" del PATH no tiene openpyxl instalado; se usó
"py -3.12" para desarrollarlo y probarlo. Si tu "python" sí trae openpyxl,
también funciona con:  python docs/generar_excel.py)

Al final, el script vuelve a abrir el archivo generado con openpyxl y
verifica (assert) que existen las 13 hojas esperadas y que los headers de
fila 1 de cada hoja de datos coinciden exactamente con los definidos aquí.
No se debe editar el resultado a mano sin volver a correr este script si se
quiere mantener sincronizado con el modelo de datos.
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Rutas
# --------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
OUTPUT_PATH = os.path.join(PROJECT_ROOT, "BASE_DATOS_GUYTON.xlsx")

# --------------------------------------------------------------------------
# Estilos
# --------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="050B2B", end_color="050B2B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(vertical="center", wrap_text=False)

# --------------------------------------------------------------------------
# Definición de las 12 hojas: (nombre_hoja, [headers], [ [fila1...], [fila2...] ], [anchos])
# El orden de las filas y de los headers debe coincidir 1:1 con docs/MODELO_DATOS.md
# --------------------------------------------------------------------------

SHEETS = []

# 1. config -----------------------------------------------------------------
SHEETS.append((
    "config",
    ["clave", "valor", "descripcion"],
    [
        ["nombre_academia", "Academia Preuniversitaria Guyton",
         "Nombre completo de la academia"],
        ["whatsapp", "+51 986 833 308",
         "Numero de WhatsApp de contacto de la academia"],
        ["drive_root_id", "",
         "ID de la carpeta raiz de Google Drive (vacio hasta que el usuario la comparta)"],
        ["color_primario", "#0829B8", "Azul real - color primario de marca"],
        ["color_acento", "#FF4A18", "Naranja Guyton - color de acento"],
        ["lema", "Asegura tu ingreso", "Lema de la academia"],
    ],
    [20, 40, 55],
))

# 2. usuarios -----------------------------------------------------------------
SHEETS.append((
    "usuarios",
    ["id_usuario", "dni", "nombres", "apellidos", "celular", "email", "rol",
     "clave_acceso", "foto_url", "estado", "fecha_registro"],
    [
        ["usr-demo-1", "70000001", "Ana", "Quispe Mamani", "986111001",
         "ana.admin@guyton.demo", "superadmin", "1111", "", "activo", "2026-06-15"],
        ["usr-demo-2", "70000002", "Carlos", "Huanca Ticona", "986111002",
         "carlos.docente@guyton.demo", "docente", "2222", "", "activo", "2026-06-15"],
        ["usr-demo-3", "70000003", "Rosa", "Flores Apaza", "986111003",
         "rosa.auxiliar@guyton.demo", "auxiliar", "3333", "", "activo", "2026-06-15"],
        ["usr-demo-4", "70000004", "Luis", "Condori Yucra", "986111004",
         "luis.estudiante@guyton.demo", "estudiante", "4444", "", "activo", "2026-06-20"],
        ["usr-demo-5", "70000005", "Maria", "Chura Pacco", "986111005",
         "maria.estudiante@guyton.demo", "estudiante", "5555", "", "activo", "2026-06-25"],
    ],
    [14, 12, 14, 18, 12, 28, 12, 14, 12, 10, 16],
))

# 3. ciclos -----------------------------------------------------------------
SHEETS.append((
    "ciclos",
    ["id_ciclo", "nombre", "anio", "fecha_inicio", "fecha_fin", "estado",
     "precio_matricula", "precio_mensualidad", "n_mensualidades", "descripcion"],
    [
        ["cic-demo-1", "2026-II", 2026, "2026-07-06", "2026-12-19", "en_curso",
         100.00, 250.00, 5, "Ciclo semestral 2026-II - DEMO"],
        ["cic-demo-2", "2027-I", 2027, "2027-01-12", "2027-06-18", "planificado",
         100.00, 250.00, 5,
         "Ciclo planificado 2027-I - DEMO (cursos aun sin docente asignado)"],
    ],
    [14, 12, 8, 14, 14, 20, 16, 18, 16, 55],
))

# 4. cursos -----------------------------------------------------------------
SHEETS.append((
    "cursos",
    ["id_curso", "nombre", "descripcion", "orden"],
    [
        ["cur-demo-1", "Matematica",
         "Algebra, aritmetica y geometria para el examen de admision - DEMO", 1],
        ["cur-demo-2", "Comunicacion",
         "Comprension lectora y redaccion - DEMO", 2],
    ],
    [12, 20, 55, 8],
))

# 5. ciclo_cursos -------------------------------------------------------------
SHEETS.append((
    "ciclo_cursos",
    ["id_ciclo_curso", "id_ciclo", "id_curso", "id_docente", "orden"],
    [
        ["cco-demo-1", "cic-demo-1", "cur-demo-1", "usr-demo-2", 1],
        ["cco-demo-2", "cic-demo-1", "cur-demo-2", "usr-demo-2", 2],
        ["cco-demo-3", "cic-demo-2", "cur-demo-1", "", 1],
    ],
    [16, 14, 12, 14, 8],
))

# 6. matriculas -----------------------------------------------------------------
SHEETS.append((
    "matriculas",
    ["id_matricula", "id_usuario", "id_ciclo", "fecha", "estado", "turno",
     "observaciones"],
    [
        ["mat-demo-1", "usr-demo-4", "cic-demo-1", "2026-06-28", "matriculado",
         "tarde", "Matricula completa - DEMO"],
        ["mat-demo-2", "usr-demo-5", "cic-demo-1", "2026-06-30", "retirado",
         "tarde",
         "DEMO - Retiro por pago de matricula rechazado y falta de regularizacion"],
    ],
    [14, 14, 14, 14, 14, 10, 55],
))

# 7. pagos -----------------------------------------------------------------
SHEETS.append((
    "pagos",
    ["id_pago", "id_matricula", "concepto", "monto", "fecha_reporte",
     "fecha_verificacion", "medio", "estado", "voucher_ref", "id_verificador"],
    [
        ["pag-demo-1", "mat-demo-1", "matricula", 100.00, "2026-06-28",
         "2026-06-29", "yape", "verificado", "YAPE-000123", "usr-demo-3"],
        ["pag-demo-2", "mat-demo-1", "mensualidad_1", 250.00, "2026-07-18",
         "", "plin", "pendiente", "PLIN-000456", ""],
        ["pag-demo-3", "mat-demo-2", "matricula", 100.00, "2026-06-30",
         "2026-07-01", "efectivo", "rechazado", "EFEC-000789", "usr-demo-3"],
    ],
    [12, 14, 16, 10, 14, 18, 12, 14, 14, 16],
))

# 8. horario -----------------------------------------------------------------
SHEETS.append((
    "horario",
    ["id_horario", "id_ciclo_curso", "dia_semana", "hora_inicio", "hora_fin",
     "aula_o_enlace"],
    [
        ["hor-demo-1", "cco-demo-1", "lunes", "18:00", "20:00",
         "Meet: meet.google.com/guyton-mat (DEMO)"],
        ["hor-demo-2", "cco-demo-2", "miercoles", "18:00", "20:00",
         "Meet: meet.google.com/guyton-com (DEMO)"],
    ],
    [14, 16, 12, 12, 10, 45],
))

# 9. clases -----------------------------------------------------------------
SHEETS.append((
    "clases",
    ["id_clase", "id_ciclo_curso", "fecha", "hora_inicio", "hora_fin", "tema",
     "modalidad", "enlace_en_vivo", "estado"],
    [
        ["cls-demo-1", "cco-demo-1", "2026-07-13", "18:00", "20:00",
         "Numeros reales y operaciones basicas (DEMO)", "virtual",
         "https://meet.google.com/guyton-mat (DEMO)", "dictada"],
        ["cls-demo-2", "cco-demo-2", "2026-07-15", "18:00", "20:00",
         "Comprension lectora: idea principal (DEMO)", "virtual",
         "https://meet.google.com/guyton-com (DEMO)", "dictada"],
        ["cls-demo-3", "cco-demo-1", "2026-07-20", "18:00", "20:00",
         "Repaso presencial: ecuaciones lineales - sesion especial (DEMO)",
         "presencial", "", "programada"],
    ],
    [12, 16, 14, 12, 10, 45, 12, 40, 14],
))

# 10. materiales -----------------------------------------------------------------
SHEETS.append((
    "materiales",
    ["id_material", "id_ciclo_curso", "id_clase", "tipo", "titulo", "semana",
     "url_drive", "id_material_padre", "fecha_publicacion", "id_autor",
     "estado"],
    [
        ["mtl-demo-1", "cco-demo-1", "cls-demo-1", "pdf_practica",
         "Practica 1: Numeros reales (DEMO)", 1,
         "https://drive.google.com/file/d/DEMO-practica-1/view", "",
         "2026-07-13", "usr-demo-2", "publicado"],
        ["mtl-demo-2", "cco-demo-1", "cls-demo-1", "pdf_resolucion",
         "Resolucion Practica 1 (DEMO)", 1,
         "https://drive.google.com/file/d/DEMO-resolucion-1/view",
         "mtl-demo-1", "2026-07-14", "usr-demo-2", "publicado"],
        ["mtl-demo-3", "cco-demo-2", "cls-demo-2", "video_grabado",
         "Grabacion: Comprension lectora - clase 1 (DEMO)", 1,
         "https://drive.google.com/file/d/DEMO-grabacion-1/view", "",
         "2026-07-15", "usr-demo-2", "borrador"],
    ],
    [14, 16, 12, 14, 40, 8, 50, 18, 18, 14, 12],
))

# 11. asistencias -----------------------------------------------------------------
SHEETS.append((
    "asistencias",
    ["id_asistencia", "id_clase", "id_usuario", "estado", "id_registrador",
     "observacion"],
    [
        ["asi-demo-1", "cls-demo-1", "usr-demo-4", "presente", "usr-demo-3",
         ""],
        ["asi-demo-2", "cls-demo-2", "usr-demo-4", "tardanza", "usr-demo-3",
         "DEMO - llego 10 minutos tarde"],
    ],
    [14, 12, 14, 12, 16, 40],
))

# 12. anuncios -----------------------------------------------------------------
SHEETS.append((
    "anuncios",
    ["id_anuncio", "id_ciclo", "titulo", "cuerpo", "fecha", "fijado",
     "id_autor", "estado"],
    [
        ["anu-demo-1", "", "Bienvenida al ciclo 2026-II (DEMO)",
         "Bienvenidos al nuevo ciclo. Revisen su horario y el grupo de WhatsApp.",
         "2026-07-01", "si", "usr-demo-1", "publicado"],
        ["anu-demo-2", "cic-demo-1", "Recordatorio: pago de mensualidad (DEMO)",
         "Recuerden reportar su pago de mensualidad antes del 25.",
         "2026-07-17", "no", "usr-demo-3", "publicado"],
    ],
    [14, 12, 35, 55, 14, 10, 14, 14],
))

# --------------------------------------------------------------------------
# Texto de la hoja LEEME
# --------------------------------------------------------------------------
LEEME_TITLE = "LEEME - BD Guyton (aula virtual, Academia Preuniversitaria Guyton)"

LEEME_LINES = [
    ("Que es este archivo", ""),
    ("", "Es la base de datos completa del aula virtual de la Academia Guyton, en formato"),
    ("", "Excel para que se pueda revisar sin depender de internet. Cuando este listo, se"),
    ("", "sube a Google Sheets y esa sera la base de datos real que use el sistema."),
    ("", ""),
    ("Como subirlo a Google Sheets", ""),
    ("", "1. Entrar a Google Drive, en la carpeta que se vaya a usar para la academia."),
    ("", "2. Subir este archivo .xlsx (arrastrarlo o usar Nuevo > Subir archivo)."),
    ("", "3. Abrirlo y elegir Archivo > Guardar como Hoja de calculo de Google."),
    ("", "   (o hacer clic derecho sobre el archivo subido > Abrir con > Google Sheets)."),
    ("", "4. Verificar que las 12 hojas de datos aparecen con sus nombres exactos y que"),
    ("", "   la fila 1 de cada una (los headers) no cambio de texto."),
    ("", ""),
    ("Advertencia importante", ""),
    ("", "NO renombrar los headers (fila 1) de ninguna hoja, ni traducirlos, ni cambiarles"),
    ("", "mayusculas/minusculas. El backend (Google Apps Script, fase 2 del proyecto) los"),
    ("", "va a referenciar tal cual estan escritos aqui. Tampoco renombrar las hojas."),
    ("", ""),
    ("Sobre los datos de este archivo", ""),
    ("", "TODAS las filas de datos en las 12 hojas son de EJEMPLO (DEMO). Se usan ids con"),
    ("", "el patron '-demo-' (ej. usr-demo-1, cic-demo-1) y la palabra DEMO en las columnas"),
    ("", "de observaciones o descripcion, precisamente para que nadie las confunda con"),
    ("", "informacion real de alumnos, docentes o pagos. Antes de operar con datos reales,"),
    ("", "estas filas deben borrarse."),
    ("", ""),
    ("Que hoja es cada una", ""),
    ("", "config       - variables globales del sistema (nombre, whatsapp, colores, etc.)"),
    ("", "usuarios     - TODAS las personas (superadmin, docentes, auxiliares, estudiantes)"),
    ("", "ciclos       - cada ciclo de preparacion (ej. 2026-II)"),
    ("", "cursos       - catalogo global de cursos (Matematica, Comunicacion, etc.)"),
    ("", "ciclo_cursos - que curso se dicta en que ciclo y con que docente (tabla puente)"),
    ("", "matriculas   - inscripcion de un usuario a un ciclo"),
    ("", "pagos        - pagos reportados/verificados, ligados a una matricula"),
    ("", "horario      - patron semanal recurrente de clases"),
    ("", "clases       - sesiones fechadas reales (donde cuelgan material y asistencia)"),
    ("", "materiales   - videos, PDFs y enlaces publicados para los alumnos"),
    ("", "asistencias  - registro de asistencia por clase y alumno"),
    ("", "anuncios     - comunicados globales o por ciclo"),
    ("", ""),
    ("Mas detalle", ""),
    ("", "El detalle completo de cada columna (tipo, obligatoriedad, ejemplos) esta en"),
    ("", "docs/MODELO_DATOS.md, y cada flujo de uso paso a paso en docs/ROLES_Y_FLUJOS.md,"),
    ("", "ambos en el repositorio del proyecto (carpeta ACADEMIA GUYTON)."),
]


def build_workbook():
    wb = Workbook()

    # --- Hoja LEEME (usa la hoja activa por defecto que crea Workbook()) ---
    ws = wb.active
    ws.title = "LEEME"
    ws.sheet_view.showGridLines = False

    ws["A1"] = LEEME_TITLE
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].fill = HEADER_FILL
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 26

    row = 3
    for label, text in LEEME_LINES:
        if label:
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = Font(bold=True, color="0829B8", size=12)
            row += 1
        if text:
            cell = ws.cell(row=row, column=1, value=text)
            cell.font = Font(size=11)
            row += 1
        if not label and not text:
            row += 1

    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 20

    # --- Las 12 hojas de datos ---
    for sheet_name, headers, rows, widths in SHEETS:
        ws = wb.create_sheet(sheet_name)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN

        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20

    return wb


def verify(path):
    """Reabre el archivo generado y verifica hojas + headers. Falla con
    AssertionError si algo no coincide con la definicion de SHEETS."""
    wb = load_workbook(path)

    expected_sheet_names = ["LEEME"] + [s[0] for s in SHEETS]
    actual_sheet_names = wb.sheetnames

    assert actual_sheet_names == expected_sheet_names, (
        f"Hojas no coinciden.\nEsperado: {expected_sheet_names}\n"
        f"Encontrado: {actual_sheet_names}"
    )

    for sheet_name, headers, rows, _widths in SHEETS:
        ws = wb[sheet_name]
        actual_headers = [ws.cell(row=1, column=i + 1).value
                           for i in range(len(headers))]
        assert actual_headers == headers, (
            f"Headers de '{sheet_name}' no coinciden.\n"
            f"Esperado: {headers}\nEncontrado: {actual_headers}"
        )

        expected_row_count = len(rows)
        actual_row_count = 0
        r = 2
        while ws.cell(row=r, column=1).value not in (None, ""):
            actual_row_count += 1
            r += 1
        assert actual_row_count == expected_row_count, (
            f"'{sheet_name}' esperaba {expected_row_count} filas de datos, "
            f"encontro {actual_row_count}"
        )

    print(f"OK - {len(expected_sheet_names)} hojas verificadas "
          f"(LEEME + {len(SHEETS)} hojas de datos) en:\n  {path}")


def main():
    wb = build_workbook()
    wb.save(OUTPUT_PATH)
    print(f"Generado: {OUTPUT_PATH}")
    verify(OUTPUT_PATH)


if __name__ == "__main__":
    main()
